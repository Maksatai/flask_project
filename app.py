from flask import Flask, render_template, request
from flask.wrappers import Request
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')
def homepage():
    excel_data_df = pd.read_excel(r'report.xlsx', engine='openpyxl')

    return render_template('index.html',excel_f=excel_data_df.values, exc=excel_data_df,count=ws.max_row)
    
outwb = load_workbook('report.xlsx')
ws = outwb.active
@app.route('/add/', methods=["POST"])
def add():
    author = request.form["author"]
    descrip = request.form["descrip"]
    year = request.form["year"]
    ws.append([author,descrip,year])
    outwb.save(filename='report.xlsx')
    outwb.close()
    return """
        <h1>Архив добавлен</h1>
        <a href='/'>Домой</a>
    """

